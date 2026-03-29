from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Depends
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import qrcode
import io
import base64
import openpyxl
import asyncio
from passlib.context import CryptContext
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Gmail SMTP setup
SENDER_EMAIL = os.getenv('SENDER_EMAIL', 'rudranishinde3101@gmail.com')
GMAIL_PASSWORD = os.getenv('GMAIL_PASSWORD', '')  # App-specific password, not regular password
GMAIL_SMTP_SERVER = 'smtp.gmail.com'
GMAIL_SMTP_PORT = 587

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============ MODELS ============

class Student(BaseModel):
    model_config = ConfigDict(extra="ignore")
    reg_no: str
    name: str
    email: EmailStr
    mobile_no: str
    dob: str
    current_year: int
    department: str = ""  # Department/Branch the student belongs to
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_till: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Faculty(BaseModel):
    model_config = ConfigDict(extra="ignore")
    faculty_id: str
    name: str
    email: EmailStr
    mobile_no: str
    department: str
    profession: str
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_till: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Visitor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    full_name: str
    phone_number: str
    email: EmailStr
    gender: str  # 'Male', 'Female', 'Other'
    date_of_birth: str
    id_type: str  # 'Aadhar Card', 'Passport', 'Driving License', etc.
    id_number: str
    id_proof_base64: Optional[str] = None
    visitor_type: str  # 'student', 'faculty', 'visitor', 'worker'
    person_to_visit_id: Optional[str] = None  # ID of student/faculty to visit
    person_to_visit_name: Optional[str] = None  # Name of person to visit
    person_to_visit_mobile: Optional[str] = None  # Mobile of person to visit
    purpose: str
    photo_base64: Optional[str] = None
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_till: str = Field(default_factory=lambda: (datetime.now(timezone.utc) + timedelta(days=1)).isoformat())
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Alumni(BaseModel):
    model_config = ConfigDict(extra="ignore")
    first_name: str
    middle_name: Optional[str] = None
    last_name: str
    email: EmailStr
    secondary_email: Optional[str] = None
    mobile_no: str
    secondary_phone: Optional[str] = None
    enrollment_number: Optional[str] = None
    gender: str  # 'Male', 'Female', 'Other'
    degree: str  # 'BE', 'BTech', 'MTech', 'PhD', 'PhDM'
    branch: str  # Branch of study
    address: str
    designation: str
    company: str
    year_of_passing: str
    year_of_joining: str
    sub_institute: str  # College/Institute name
    registered_for_portal: str  # 'yes' or 'no'
    linkedin_profile: Optional[str] = None
    whom_to_meet: str  # 'faculty', 'student', 'visitor'
    selected_person_id: Optional[str] = None
    selected_person_name: Optional[str] = None
    selected_person_mobile: Optional[str] = None
    photo_base64: Optional[str] = None
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_till: str = Field(default_factory=lambda: (datetime.now(timezone.utc) + timedelta(days=1)).isoformat())
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class FacultyRegisteredVisitor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    visitor_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    faculty_id: str  # The faculty who registered this visitor
    visitor_type: str  # 'HR', 'Relative', 'Other Faculty', 'Student', 'Other'
    name: str
    email: EmailStr
    mobile_no: str
    date_from: str  # Date visitor is allowed from
    date_to: str    # Date visitor is allowed until
    valid_till: str  # QR code validity
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Event(BaseModel):
    model_config = ConfigDict(extra="ignore")
    event_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_name: str
    event_type: str
    date_from: str
    date_to: str
    description: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class EventStudent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    event_id: str
    reg_no: str
    name: str
    email: EmailStr
    mobile_no: str
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    valid_from: str
    valid_to: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Request models
class AdminLogin(BaseModel):
    username: str
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    mobile_no: str

class GuardLogin(BaseModel):
    username: str
    password: str

class StudentCreate(BaseModel):
    reg_no: str
    name: str
    email: EmailStr
    mobile_no: str
    dob: str
    current_year: int
    department: str = ""  # Department/Branch

class FacultyCreate(BaseModel):
    name: str
    email: EmailStr
    mobile_no: str
    department: str
    profession: str
    valid_till: str

class EventCreate(BaseModel):
    event_name: str
    event_type: str
    date_from: str
    date_to: str
    description: str

class EventStudentCreate(BaseModel):
    event_id: str
    reg_no: str

class QRValidate(BaseModel):
    token: str

# ============ HELPER FUNCTIONS ============

def calculate_student_validity(current_year: int) -> str:
    """Calculate validity based on current year"""
    today = datetime.now(timezone.utc)
    years_remaining = 4 - current_year
    if years_remaining <= 0:
        return (today + timedelta(days=365)).isoformat()
    valid_till = today + timedelta(days=365 * years_remaining)
    return valid_till.isoformat()

async def get_next_faculty_id() -> str:
    """Generate next faculty ID in format PICT-FAC-001"""
    # Find the highest faculty_id number across all records
    result = await db.faculty.find_one(
        {},
        sort=[("faculty_id", -1)],
        projection={"_id": 0, "faculty_id": 1}
    )
    if not result:
        return "PICT-FAC-001"
    last_id = result.get("faculty_id", "PICT-FAC-000")
    try:
        # Handle both old (fact_XXX) and new (PICT-FAC-XXX) formats
        if "PICT-FAC-" in last_id:
            num = int(last_id.split("-")[-1]) + 1
        else:
            num = int(last_id.split("_")[1]) + 1
    except (ValueError, IndexError):
        num = 1
    return f"PICT-FAC-{num:03d}"

async def generate_external_student_id(event_name: str, event_id: str) -> str:
    """Generate unique ID for external students (EventName_exe001 format)"""
    # Sanitize event name: remove spaces, special chars, keep lettters/numbers/underscores only
    sanitized_name = "".join(c if c.isalnum() or c == "_" else "_" for c in event_name)
    # Replace multiple underscores with single underscore and remove trailing/leading
    sanitized_name = "_".join(filter(None, sanitized_name.split("_")))
    
    # Get count of external students already added for this event
    count = await db.event_students.count_documents({"event_id": event_id})
    # Start from 1 (001 format)
    next_num = count + 1
    
    # Format: EventName_exe001
    return f"{sanitized_name}_exe{next_num:03d}"


    """Generate QR code as base64 string"""
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(token)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode()

async def send_qr_email(recipient_email: str, name: str, token: str, valid_till: str) -> dict:
    """Send QR code via Gmail SMTP - Returns dict with status"""
    # Validate recipient email
    if not recipient_email or "@" not in recipient_email:
        return {
            "success": False, 
            "email": recipient_email,
            "reason": f"Invalid email format: {recipient_email}"
        }
    
    if not GMAIL_PASSWORD:
        logger.warning("Gmail password not configured in .env file")
        return {
            "success": False,
            "email": recipient_email,
            "reason": "Gmail not configured (.env missing GMAIL_PASSWORD)"
        }
    
    qr_base64 = generate_qr_code(token)
    
    html_content = f"""
    <html>
        <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
            <div style="max-width: 600px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <h2 style="color: #333; border-bottom: 3px solid #007bff; padding-bottom: 10px;">PICT Guard - Your Gate Pass</h2>
                <p style="color: #555; font-size: 16px;">Hello {name},</p>
                <p style="color: #555; font-size: 14px;">Your gate pass has been generated successfully. Please find your QR code below:</p>
                <p style="color: #888; font-size: 12px;"><strong>Valid Until:</strong> {valid_till}</p>
                <div style="text-align: center; margin: 30px 0; padding: 20px; background-color: #f9f9f9; border-radius: 5px;">
                    <img src="cid:qrcode" alt="QR Code" style="max-width: 300px; border: 2px solid #ddd; padding: 10px;"/>
                </div>
                <p style="color: #555; font-size: 14px;"><strong>Instructions:</strong> Please show this QR code at the gate for entry.</p>
                <hr style="border: none; border-top: 1px solid #ddd; margin: 20px 0;">
                <p style="color: #888; font-size: 12px;">Best regards,<br><strong>PICT Guard Team</strong></p>
            </div>
        </body>
    </html>
    """
    
    try:
        logger.info(f"Attempting to send email from {SENDER_EMAIL} to {recipient_email}")
        
        # Create message
        msg = MIMEMultipart('related')
        msg['Subject'] = "PICT Guard - Your Gate Pass QR Code"
        msg['From'] = SENDER_EMAIL
        msg['To'] = recipient_email
        
        # Attach HTML content
        msg_alternative = MIMEMultipart('alternative')
        msg.attach(msg_alternative)
        msg_alternative.attach(MIMEText(html_content, 'html'))
        
        # Attach QR code image
        qr_image = MIMEImage(base64.b64decode(qr_base64), name='qrcode.png')
        qr_image.add_header('Content-ID', '<qrcode>')
        qr_image.add_header('Content-Disposition', 'inline', filename='qrcode.png')
        msg.attach(qr_image)
        
        # Send email via Gmail SMTP with timeout
        def send_email():
            try:
                with smtplib.SMTP(GMAIL_SMTP_SERVER, GMAIL_SMTP_PORT, timeout=10) as server:
                    server.starttls()
                    server.login(SENDER_EMAIL, GMAIL_PASSWORD)
                    server.send_message(msg)
                    return True
            except smtplib.SMTPAuthenticationError as e:
                logger.error(f"Gmail Authentication Error - Check SENDER_EMAIL and GMAIL_PASSWORD: {str(e)}")
                raise
            except smtplib.SMTPException as e:
                logger.error(f"Gmail SMTP Error: {str(e)}")
                raise
            except TimeoutError as e:
                logger.error(f"Gmail Connection Timeout: Could not connect to {GMAIL_SMTP_SERVER}:{GMAIL_SMTP_PORT} - {str(e)}")
                raise
            except Exception as e:
                logger.error(f"Unexpected Email Error: {str(e)}")
                raise
        
        await asyncio.to_thread(send_email)
        logger.info(f"Email sent successfully to {recipient_email}")
        return {
            "success": True,
            "email": recipient_email,
            "reason": "Email sent successfully"
        }
    except smtplib.SMTPAuthenticationError as e:
        error_msg = f"Gmail authentication failed for {SENDER_EMAIL}. Check your GMAIL_PASSWORD in .env file."
        logger.error(error_msg)
        return {
            "success": False,
            "email": recipient_email,
            "reason": error_msg
        }
    except TimeoutError as e:
        error_msg = f"Gmail server connection timeout. Check your internet connection."
        logger.error(f"Timeout sending email to {recipient_email}: {str(e)}")
        return {
            "success": False,
            "email": recipient_email,
            "reason": error_msg
        }
    except Exception as e:
        error_msg = f"Failed to send email to {recipient_email}: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "email": recipient_email,
            "reason": str(e)
        }


# ============ AUTH ROUTES ============

@api_router.post("/auth/admin/login")
async def admin_login(data: AdminLogin):
    if data.username == "admin" and data.password == "admin123":
        return {"success": True, "role": "admin", "token": str(uuid.uuid4())}
    raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.post("/auth/user/login")
async def user_login(data: UserLogin):
    # Check student
    student = await db.students.find_one(
        {"email": data.email, "mobile_no": data.mobile_no},
        {"_id": 0}
    )
    if student:
        return {"success": True, "role": "student", "data": student}
    
    # Check faculty
    faculty = await db.faculty.find_one(
        {"email": data.email, "mobile_no": data.mobile_no},
        {"_id": 0}
    )
    if faculty:
        return {"success": True, "role": "faculty", "data": faculty}
    
    raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.post("/auth/guard/login")
async def guard_login(data: GuardLogin):
    # Simple guard login - username can be guard1, guard2, etc., password is "guard123"
    if data.username.startswith("guard") and data.password == "guard123":
        return {"success": True, "role": "guard", "guard_id": data.username}
    raise HTTPException(status_code=401, detail="Invalid credentials")

# ============ STUDENT ROUTES ============

@api_router.post("/students", response_model=Student)
async def create_student(data: StudentCreate):
    # Check if student exists by reg_no, email, or mobile_no
    existing = await db.students.find_one(
        {"$or": [
            {"reg_no": data.reg_no},
            {"email": data.email},
            {"mobile_no": data.mobile_no}
        ]},
        {"_id": 0}
    )
    if existing:
        if existing.get('reg_no') == data.reg_no:
            raise HTTPException(status_code=400, detail=f"Student with Reg No '{data.reg_no}' already exists")
        elif existing.get('email') == data.email:
            raise HTTPException(status_code=400, detail=f"Student with Email '{data.email}' already exists")
        else:
            raise HTTPException(status_code=400, detail=f"Student with Mobile '{data.mobile_no}' already exists")
    
    valid_till = calculate_student_validity(data.current_year)
    student = Student(**data.model_dump(), valid_till=valid_till)
    await db.students.insert_one(student.model_dump())
    return student

@api_router.post("/students/bulk")
async def bulk_create_students(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files allowed")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    students_created = 0
    duplicates = []
    errors = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row or not row[0]:  # Skip empty rows
                continue
            
            # Excel format: RegNo, Name, Email, Mobile, DOB, Year, Department (7 columns, Department is optional)
            actual_rows = row
            
            # Validate we have the required 6 columns (Department is optional)
            if len(actual_rows) < 6:
                errors.append(f"Row {idx}: Missing required columns (need at least 6: RegNo, Name, Email, Mobile, DOB, Year)")
                logger.error(f"Row {idx} column count: {len(actual_rows)}")
                continue
            
            # Check for None/empty values (first 6 are required, 7th is optional)
            if not actual_rows[0] or not actual_rows[1] or not actual_rows[2] or not actual_rows[3] or not actual_rows[4] or not actual_rows[5]:
                missing_fields = []
                field_names = ["RegNo", "Name", "Email", "Mobile", "DOB", "Year"]
                for i, val in enumerate(actual_rows[:6]):
                    if not val:
                        missing_fields.append(field_names[i])
                errors.append(f"Row {idx}: Missing required field(s): {', '.join(missing_fields)}")
                continue
            
            try:
                # Validate Year field specifically - must be integer 1-4
                try:
                    year_val = int(actual_rows[5])
                    if year_val < 1 or year_val > 4:
                        errors.append(f"Row {idx}: Year must be 1, 2, 3, or 4. Got: {actual_rows[5]}")
                        continue
                except ValueError:
                    errors.append(f"Row {idx}: Year field (column 6) must be a number (1-4), not '{actual_rows[5]}'")
                    continue
                
                # Department is optional (column 7)
                department = ""
                if len(actual_rows) >= 7 and actual_rows[6]:
                    department = str(actual_rows[6]).strip()
                
                data = StudentCreate(
                    reg_no=str(actual_rows[0]).strip(),
                    name=str(actual_rows[1]).strip(),
                    email=str(actual_rows[2]).strip(),
                    mobile_no=str(actual_rows[3]).strip(),
                    dob=str(actual_rows[4]).strip(),
                    current_year=year_val,
                    department=department
                )
            except Exception as validation_error:
                errors.append(f"Row {idx}: Invalid data format - {str(validation_error)}")
                logger.error(f"Row {idx} validation error: {str(validation_error)}")
                continue
            
            # Check if exists by reg_no, email, or mobile_no
            existing = await db.students.find_one(
                {"$or": [
                    {"reg_no": data.reg_no},
                    {"email": data.email},
                    {"mobile_no": data.mobile_no}
                ]},
                {"_id": 0}
            )
            if existing:
                # Remove the old duplicate and add the new one
                await db.students.delete_one({
                    "$or": [
                        {"reg_no": data.reg_no},
                        {"email": data.email},
                        {"mobile_no": data.mobile_no}
                    ]
                })
                duplicates.append(f"Row {idx}: {data.reg_no} (Removed old duplicate, adding updated record)")
            
            # Use the reg_no directly from Excel (no auto-generation)
            valid_till = calculate_student_validity(data.current_year)
            student = Student(**data.model_dump(), valid_till=valid_till)
            await db.students.insert_one(student.model_dump())
            students_created += 1
        except Exception as e:
            error_msg = f"Row {idx}: {str(e)}"
            errors.append(error_msg)
            logger.error(error_msg)
    
    return {
        "success": True, 
        "created": students_created, 
        "duplicates": len(duplicates), 
        "duplicate_details": duplicates,
        "error_count": len(errors),
        "error_details": errors  # Show detailed error list
    }

@api_router.get("/students", response_model=List[Student])
async def get_students():
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    return students

# ============ FACULTY ROUTES ============

@api_router.post("/faculty", response_model=Faculty)
async def create_faculty(data: FacultyCreate):
    # Check if faculty exists by email or mobile_no
    existing = await db.faculty.find_one(
        {"$or": [
            {"email": data.email},
            {"mobile_no": data.mobile_no}
        ]},
        {"_id": 0}
    )
    if existing:
        if existing.get('email') == data.email:
            raise HTTPException(status_code=400, detail=f"Faculty with Email '{data.email}' already exists")
        else:
            raise HTTPException(status_code=400, detail=f"Faculty with Mobile '{data.mobile_no}' already exists")
    
    faculty_id = await get_next_faculty_id()
    faculty = Faculty(**data.model_dump(), faculty_id=faculty_id)
    await db.faculty.insert_one(faculty.model_dump())
    return faculty

@api_router.post("/faculty/bulk")
async def bulk_create_faculty(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files allowed")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    faculty_created = 0
    duplicates = []
    errors = []
    
    # Get the highest existing faculty_id to ensure no duplicates in this batch
    last_result = await db.faculty.find_one(
        {},
        sort=[("faculty_id", -1)],
        projection={"faculty_id": 1}
    )
    try:
        if last_result and "PICT-FAC-" in last_result["faculty_id"]:
            last_num = int(last_result["faculty_id"].split("-")[-1])
        elif last_result and "_" in last_result["faculty_id"]:
            last_num = int(last_result["faculty_id"].split("_")[1])
        else:
            last_num = 0
    except (ValueError, IndexError, TypeError):
        last_num = 0
    
    next_id_num = last_num + 1
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row or not row[0]:  # Skip empty rows
                continue
            
            # Handle both cases: with and without Faculty ID column
            # If 7 columns, assume first is Faculty ID (auto-generated, ignore it)
            # If 6 columns, use all as: Name, Email, Mobile, Department, Profession, Valid Till
            col_offset = 0
            if len(row) >= 7 and row[0]:  # Might have Faculty ID in first column
                col_offset = 1
                actual_rows = row[1:]
            else:
                actual_rows = row
            
            # Validate we have the required 6 columns
            if len(actual_rows) < 6:
                errors.append(f"Row {idx}: Missing required columns (need 6: Name, Email, Mobile, Department, Profession, Valid Till)")
                logger.error(f"Row {idx} column count: {len(actual_rows)}, raw row length: {len(row)}")
                continue
            
            # Check for None/empty values
            if not actual_rows[0] or not actual_rows[1] or not actual_rows[2] or not actual_rows[3] or not actual_rows[4] or not actual_rows[5]:
                missing_fields = []
                field_names = ["Name", "Email", "Mobile", "Department", "Profession", "Valid Till"]
                for i, val in enumerate(actual_rows[:6]):
                    if not val:
                        missing_fields.append(field_names[i])
                errors.append(f"Row {idx}: Missing required field(s): {', '.join(missing_fields)}")
                continue
            
            try:
                data = FacultyCreate(
                    name=str(actual_rows[0]).strip(),
                    email=str(actual_rows[1]).strip(),
                    mobile_no=str(actual_rows[2]).strip(),
                    department=str(actual_rows[3]).strip(),
                    profession=str(actual_rows[4]).strip(),
                    valid_till=str(actual_rows[5]).strip()
                )
            except Exception as validation_error:
                errors.append(f"Row {idx}: Invalid data format - {str(validation_error)}")
                logger.error(f"Row {idx} validation error: {str(validation_error)}")
                continue
            
            # Check if exists by email or mobile_no
            existing = await db.faculty.find_one(
                {"$or": [
                    {"email": data.email},
                    {"mobile_no": data.mobile_no}
                ]},
                {"_id": 0}
            )
            if existing:
                # Remove the old duplicate and add the new one
                await db.faculty.delete_one({
                    "$or": [
                        {"email": data.email},
                        {"mobile_no": data.mobile_no}
                    ]
                })
                duplicates.append(f"Row {idx}: {data.name} (Removed old duplicate, adding updated record)")
            
            # Use the local counter to generate unique IDs
            faculty_id = f"PICT-FAC-{next_id_num:03d}"
            next_id_num += 1
            faculty = Faculty(**data.model_dump(), faculty_id=faculty_id)
            await db.faculty.insert_one(faculty.model_dump())
            faculty_created += 1
        except Exception as e:
            error_msg = f"Row {idx}: {str(e)}"
            errors.append(error_msg)
            logger.error(error_msg)
    
    return {
        "success": True, 
        "created": faculty_created, 
        "duplicates": len(duplicates), 
        "duplicate_details": duplicates, 
        "error_count": len(errors),
        "error_details": errors  # Show detailed error list
    }

@api_router.get("/faculty", response_model=List[Faculty])
async def get_faculty():
    faculty = await db.faculty.find({}, {"_id": 0}).to_list(1000)
    return faculty

# ============ EVENT ROUTES ============

@api_router.post("/events", response_model=Event)
async def create_event(data: EventCreate):
    event = Event(**data.model_dump())
    await db.events.insert_one(event.model_dump())
    return event

@api_router.get("/events", response_model=List[Event])
async def get_events():
    events = await db.events.find({}, {"_id": 0}).to_list(1000)
    return events

class EventStudentManualCreate(BaseModel):
    event_id: str
    reg_no: Optional[str] = None  # Optional - auto-generated if not provided
    name: str
    email: EmailStr
    mobile_no: str

@api_router.post("/events/students")
async def add_student_to_event(data: EventStudentCreate):
    # Get student details
    student = await db.students.find_one({"reg_no": data.reg_no}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found in database. Use manual entry for external students.")
    
    # Get event details
    event = await db.events.find_one({"event_id": data.event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Create event student entry
    event_student = EventStudent(
        event_id=data.event_id,
        reg_no=student["reg_no"],
        name=student["name"],
        email=student["email"],
        mobile_no=student["mobile_no"],
        valid_from=event["date_from"],
        valid_to=event["date_to"]
    )
    
    await db.event_students.insert_one(event_student.model_dump())
    
    # Send QR email with error handling
    email_result = await send_qr_email(
        event_student.email,
        event_student.name,
        event_student.token,
        event_student.valid_to
    )
    
    response_msg = "Student added to event"
    if email_result["success"]:
        response_msg += " and email sent successfully"
    else:
        response_msg += f" (Email failed: {email_result['reason']})"
    
    return {"success": True, "message": response_msg}

@api_router.post("/events/students/manual")
async def add_manual_student_to_event(data: EventStudentManualCreate):
    # Get event details
    event = await db.events.find_one({"event_id": data.event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Check if student with same email/mobile already exists in this event
    existing_student = await db.event_students.find_one(
        {
            "event_id": data.event_id,
            "$or": [
                {"email": data.email},
                {"mobile_no": data.mobile_no}
            ]
        },
        {"_id": 0}
    )
    
    if existing_student:
        # Student already in event - return existing record with reused ID
        return {
            "success": True,
            "message": f"Student already registered for this event with ID: {existing_student['reg_no']}",
            "student_id": existing_student['reg_no'],
            "reused": True
        }
    
    # Auto-generate reg_no if not provided (for external students)
    if data.reg_no:
        reg_no = data.reg_no
    else:
        reg_no = await generate_external_student_id(event["event_name"], data.event_id)
    
    # Create event student entry with manual data
    event_student = EventStudent(
        event_id=data.event_id,
        reg_no=reg_no,
        name=data.name,
        email=data.email,
        mobile_no=data.mobile_no,
        valid_from=event["date_from"],
        valid_to=event["date_to"]
    )
    
    await db.event_students.insert_one(event_student.model_dump())
    
    # Send QR email with error handling
    email_result = await send_qr_email(
        event_student.email,
        event_student.name,
        event_student.token,
        event_student.valid_to
    )
    
    response_msg = f"External student added to event (ID: {reg_no})"
    if email_result["success"]:
        response_msg += " and email sent successfully"
    else:
        response_msg += f" (Email failed: {email_result['reason']})"
    
    return {"success": True, "message": response_msg, "student_id": reg_no}

@api_router.post("/events/students/bulk")
async def bulk_add_students_to_event(event_id: str = Form(...), file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files allowed")
    
    # Get event details
    event = await db.events.find_one({"event_id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    students_added = 0
    duplicates = []
    errors = []
    email_errors = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row or (not row[0] and not row[1]):  # Skip completely empty rows
                continue
            
            email = None
            reg_no = None
            
            # Check if it's a manual entry (has name, email, mobile in Excel)
            if len(row) >= 4 and row[1] and row[2] and row[3]:
                # Manual entry for external students (name, email, mobile provided)
                try:
                    name = str(row[1]).strip()
                    email = str(row[2]).strip()
                    mobile_no = str(row[3]).strip()
                    
                    # Check if student with same email/mobile already exists in this event
                    existing_student = await db.event_students.find_one(
                        {
                            "event_id": event_id,
                            "$or": [
                                {"email": email},
                                {"mobile_no": mobile_no}
                            ]
                        },
                        {"_id": 0}
                    )
                    
                    if existing_student:
                        duplicates.append(f"Row {idx}: {name} ({email}) - Already registered with ID: {existing_student['reg_no']}")
                        continue
                    
                    # Auto-generate reg_no if first column is empty (for external students)
                    if row[0]:
                        reg_no = str(row[0]).strip()
                    else:
                        reg_no = await generate_external_student_id(event["event_name"], event_id)
                    
                    event_student = EventStudent(
                        event_id=event_id,
                        reg_no=reg_no,
                        name=name,
                        email=email,
                        mobile_no=mobile_no,
                        valid_from=event["date_from"],
                        valid_to=event["date_to"]
                    )
                except Exception as validation_error:
                    errors.append(f"Row {idx}: Invalid manual entry format - {str(validation_error)}")
                    logger.error(f"Row {idx} validation error: {str(validation_error)}")
                    continue
            else:
                # Try to find in existing students (only reg_no provided)
                reg_no = str(row[0]).strip() if row[0] else None
                
                if not reg_no:
                    errors.append(f"Row {idx}: Missing registration number or complete student details")
                    continue
                
                student = await db.students.find_one({"reg_no": reg_no}, {"_id": 0})
                
                if not student:
                    errors.append(f"Row {idx}: PICT Student {reg_no} not found in database. Provide name, email, mobile for external students")
                    continue
                
                # Check if already added to this event
                existing_event_student = await db.event_students.find_one(
                    {
                        "event_id": event_id,
                        "email": student["email"]
                    },
                    {"_id": 0}
                )
                
                if existing_event_student:
                    duplicates.append(f"Row {idx}: {student['name']} ({reg_no}) - Already registered for this event")
                    continue
                
                try:
                    event_student = EventStudent(
                        event_id=event_id,
                        reg_no=student["reg_no"],
                        name=student["name"],
                        email=student["email"],
                        mobile_no=student["mobile_no"],
                        valid_from=event["date_from"],
                        valid_to=event["date_to"]
                    )
                    email = student["email"]
                except Exception as validation_error:
                    errors.append(f"Row {idx}: Error creating event student - {str(validation_error)}")
                    logger.error(f"Row {idx} validation error: {str(validation_error)}")
                    continue
            
            await db.event_students.insert_one(event_student.model_dump())
            
            # Send email with error handling - continue if email fails
            email_result = await send_qr_email(
                event_student.email,
                event_student.name,
                event_student.token,
                event_student.valid_to
            )
            
            if not email_result["success"]:
                # Log email failure but continue processing
                email_errors.append({
                    "row": idx,
                    "name": event_student.name,
                    "email": event_student.email,
                    "student_id": event_student.reg_no,
                    "reason": email_result["reason"]
                })
                logger.warning(f"Row {idx}: Email failed for {event_student.name} ({event_student.email}): {email_result['reason']}")
            
            students_added += 1
        except Exception as e:
            error_msg = f"Row {idx}: {str(e)}"
            errors.append(error_msg)
            logger.error(error_msg)
    
    return {
        "success": True, 
        "added": students_added, 
        "duplicates": len(duplicates), 
        "duplicate_details": duplicates,
        "error_count": len(errors),
        "error_details": errors,
        "email_errors": len(email_errors),
        "email_error_details": email_errors
    }

@api_router.get("/events/{event_id}/students", response_model=List[EventStudent])
async def get_event_students(event_id: str):
    students = await db.event_students.find({"event_id": event_id}, {"_id": 0}).to_list(1000)
    return students

@api_router.post("/events/{event_id}/students/{token}/resend-email")
async def resend_student_email(event_id: str, token: str):
    """Resend QR code email to a specific event student"""
    student = await db.event_students.find_one(
        {"event_id": event_id, "token": token},
        {"_id": 0}
    )
    
    if not student:
        raise HTTPException(status_code=404, detail="Student not found in event")
    
    # Resend QR email
    email_result = await send_qr_email(
        student["email"],
        student["name"],
        student["token"],
        student["valid_to"]
    )
    
    if not email_result["success"]:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to send email: {email_result['reason']}"
        )
    
    return {
        "success": True,
        "message": f"Email resent successfully to {student['email']}",
        "student": {
            "name": student["name"],
            "email": student["email"],
            "reg_no": student["reg_no"]
        }
    }

@api_router.post("/events/{event_id}/bulk-resend-emails")
async def bulk_resend_emails(event_id: str):
    """Resend QR codes to all students in an event"""
    event = await db.events.find_one({"event_id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    students = await db.event_students.find({"event_id": event_id}, {"_id": 0}).to_list(1000)
    
    if not students:
        raise HTTPException(status_code=404, detail="No students found for this event")
    
    success_count = 0
    failed_emails = []
    
    for student in students:
        email_result = await send_qr_email(
            student["email"],
            student["name"],
            student["token"],
            student["valid_to"]
        )
        
        if email_result["success"]:
            success_count += 1
        else:
            failed_emails.append({
                "name": student["name"],
                "email": student["email"],
                "reason": email_result["reason"]
            })
    
    return {
        "success": True,
        "message": f"Emails sent to {success_count}/{len(students)} students",
        "sent": success_count,
        "failed": len(failed_emails),
        "failed_details": failed_emails
    }

# ============ SEARCH ROUTES FOR ALUMNI ============

@api_router.get("/search/faculty")
async def search_faculty(query: str):
    """Search faculty by name or mobile number"""
    search_filter = {
        "$or": [
            {"name": {"$regex": query, "$options": "i"}},
            {"mobile_no": {"$regex": query, "$options": "i"}}
        ]
    }
    results = await db.faculty.find(search_filter, {"_id": 0}).to_list(100)
    return results

@api_router.get("/search/students")
async def search_students(query: str):
    """Search students by name, mobile number, or reg_no (GRN)"""
    search_filter = {
        "$or": [
            {"name": {"$regex": query, "$options": "i"}},
            {"mobile_no": {"$regex": query, "$options": "i"}},
            {"reg_no": {"$regex": query, "$options": "i"}}
        ]
    }
    results = await db.students.find(search_filter, {"_id": 0}).to_list(100)
    return results

# ============ VISITOR & ALUMNI ROUTES ============

@api_router.post("/visitors", response_model=Visitor)
async def register_visitor(data: Visitor):
    await db.visitors.insert_one(data.model_dump())
    
    # Send QR email with full_name
    await send_qr_email(data.email, data.full_name, data.token, data.valid_till)
    
    return data

@api_router.get("/visitors")
async def get_visitors():
    try:
        visitors = await db.visitors.find({}, {"_id": 0}).to_list(1000)
        return visitors if visitors else []
    except Exception as e:
        logger.error(f"Error fetching visitors: {str(e)}")
        return []

@api_router.get("/visitors/meeting/{person_id}")
async def get_visitor_meetings(person_id: str):
    """Get all visitors coming to meet a specific student/faculty"""
    try:
        visitors = await db.visitors.find(
            {"person_to_visit_id": person_id},
            {"_id": 0}
        ).to_list(1000)
        return visitors if visitors else []
    except Exception as e:
        logger.error(f"Error fetching visitor meetings: {str(e)}")
        return []

@api_router.get("/alumni/meeting/{person_id}")
async def get_alumni_meetings(person_id: str):
    """Get all alumni coming to meet a specific student/faculty"""
    try:
        alumni = await db.alumni.find(
            {"selected_person_id": person_id},
            {"_id": 0}
        ).to_list(1000)
        return alumni if alumni else []
    except Exception as e:
        logger.error(f"Error fetching alumni meetings: {str(e)}")
        return []

@api_router.post("/alumni", response_model=Alumni)
async def register_alumni(data: Alumni):
    await db.alumni.insert_one(data.model_dump())
    
    # Construct full name from first_name and last_name
    full_name = f"{data.first_name} {data.last_name}".strip()
    
    # Send QR email
    await send_qr_email(data.email, full_name, data.token, data.valid_till)
    
    return data

@api_router.get("/alumni")
async def get_alumni():
    try:
        alumni = await db.alumni.find({}, {"_id": 0}).to_list(1000)
        return alumni if alumni else []
    except Exception as e:
        logger.error(f"Error fetching alumni: {str(e)}")
        return []

@api_router.get("/alumni/meeting/{person_id}")
async def get_alumni_meetings(person_id: str):
    """Get all alumni coming to meet a specific student/faculty"""
    try:
        alumni = await db.alumni.find(
            {"selected_person_id": person_id},
            {"_id": 0}
        ).to_list(1000)
        return alumni if alumni else []
    except Exception as e:
        logger.error(f"Error fetching alumni meetings: {str(e)}")
        return []

# ============ FACULTY REGISTERED VISITORS ============

@api_router.post("/faculty/register-visitor", response_model=FacultyRegisteredVisitor)
async def faculty_register_visitor(data: FacultyRegisteredVisitor):
    """Faculty registers a visitor (HR, Relative, Other Faculty, Student, Other)"""
    try:
        # Check if visitor already registered by this faculty (by email or mobile)
        existing_visitor = await db.faculty_registered_visitors.find_one(
            {
                "faculty_id": data.faculty_id,
                "$or": [
                    {"email": data.email},
                    {"mobile_no": data.mobile_no}
                ]
            },
            {"_id": 0}
        )
        
        if existing_visitor:
            if existing_visitor["email"] == data.email:
                raise HTTPException(
                    status_code=400,
                    detail=f"Visitor with email '{data.email}' is already registered"
                )
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Visitor with mobile '{data.mobile_no}' is already registered"
                )
        
        visitor_doc = data.model_dump()
        await db.faculty_registered_visitors.insert_one(visitor_doc)
        
        # Send QR email to the visitor
        full_name = data.name
        await send_qr_email(data.email, full_name, data.token, data.valid_till)
        
        logger.info(f"Faculty {data.faculty_id} registered visitor: {data.name}")
        return data
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error registering faculty visitor: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error registering visitor: {str(e)}")

@api_router.get("/faculty/{faculty_id}/registered-visitors")
async def get_faculty_registered_visitors(faculty_id: str):
    """Get all visitors registered by a specific faculty"""
    try:
        visitors = await db.faculty_registered_visitors.find(
            {"faculty_id": faculty_id},
            {"_id": 0}
        ).to_list(1000)
        return visitors if visitors else []
    except Exception as e:
        logger.error(f"Error fetching faculty registered visitors: {str(e)}")
        return []

@api_router.post("/faculty/visitor/resend-email")
async def resend_visitor_qr_email(data: dict):
    """Resend QR email to a visitor"""
    try:
        visitor_id = data.get("visitor_id")
        
        # Find the visitor
        visitor = await db.faculty_registered_visitors.find_one(
            {"visitor_id": visitor_id},
            {"_id": 0}
        )
        
        if not visitor:
            raise HTTPException(status_code=404, detail="Visitor not found")
        
        # Send email
        success = await send_qr_email(visitor["email"], visitor["name"], visitor["token"], visitor["valid_till"])
        
        if success:
            return {"success": True, "message": f"Email resent to {visitor['email']}"}
        else:
            raise HTTPException(status_code=500, detail="Failed to send email")
    except Exception as e:
        logger.error(f"Error resending email: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")

# ============ QR VALIDATION ============

@api_router.post("/validate-qr")
async def validate_qr(data: QRValidate):
    token = data.token
    now = datetime.now(timezone.utc)
    
    # Check all collections
    collections = [
        ("students", "Student"),
        ("faculty", "Faculty"),
        ("visitors", "Visitor"),
        ("alumni", "Alumni"),
        ("event_students", "Event Student"),
        ("faculty_registered_visitors", "Registered Visitor")
    ]
    
    for collection_name, user_type in collections:
        try:
            collection = db[collection_name]
            record = await collection.find_one({"token": token}, {"_id": 0})
            
            # If record not found in this collection, try next one
            if not record:
                continue
            
            # Determine which field to use for validity check
            if collection_name == "event_students":
                valid_till_str = record.get("valid_to")
            else:
                valid_till_str = record.get("valid_till")
            
            # If no validity field, skip this record
            if not valid_till_str:
                logger.warning(f"No validity field found for {collection_name} with token {token}")
                continue
            
            # Parse the datetime safely
            try:
                valid_till = datetime.fromisoformat(valid_till_str.replace("Z", "+00:00"))
            except Exception as e:
                logger.error(f"Error parsing datetime for {collection_name}: {str(e)}")
                continue
            
            # Check if token is valid
            if now <= valid_till:
                logger.info(f"Token validation successful for {user_type}")
                return {
                    "valid": True,
                    "type": user_type,
                    "name": record.get("name"),
                    "email": record.get("email"),
                    "valid_till": valid_till_str,
                    "details": record
                }
            else:
                # Token expired
                logger.info(f"Token expired for {user_type}")
                return {
                    "valid": False,
                    "reason": "QR code expired",
                    "type": user_type,
                    "name": record.get("name"),
                    "expired_on": valid_till_str
                }
        except Exception as e:
            logger.error(f"Error checking {collection_name}: {str(e)}")
            continue
    
    logger.warning(f"Token not found: {token}")
    return {"valid": False, "reason": "QR code not found"}

# ============ DEBUG ENDPOINT ============

@api_router.get("/debug/token/{token}")
async def debug_token(token: str):
    """Debug endpoint to check if token exists in any collection"""
    results = {
        "token": token,
        "found_in": [],
        "collections_checked": []
    }
    
    collections = [
        "students",
        "faculty",
        "visitors",
        "alumni",
        "event_students",
        "faculty_registered_visitors"
    ]
    
    for collection_name in collections:
        try:
            collection = db[collection_name]
            record = await collection.find_one({"token": token}, {"_id": 0})
            
            results["collections_checked"].append(collection_name)
            
            if record:
                result_entry = {
                    "collection": collection_name,
                    "name": record.get("name"),
                    "email": record.get("email"),
                }
                
                # Check validity fields
                if "valid_till" in record:
                    result_entry["validity_field"] = "valid_till"
                    result_entry["valid_till"] = record.get("valid_till")
                    result_entry["current_time"] = datetime.now(timezone.utc).isoformat()
                    try:
                        valid_till = datetime.fromisoformat(record.get("valid_till").replace("Z", "+00:00"))
                        result_entry["is_valid"] = datetime.now(timezone.utc) <= valid_till
                    except Exception as e:
                        result_entry["is_valid"] = f"Error parsing: {str(e)}"
                
                if "valid_to" in record:
                    result_entry["validity_field"] = "valid_to"
                    result_entry["valid_to"] = record.get("valid_to")
                    result_entry["current_time"] = datetime.now(timezone.utc).isoformat()
                    try:
                        valid_to = datetime.fromisoformat(record.get("valid_to").replace("Z", "+00:00"))
                        result_entry["is_valid"] = datetime.now(timezone.utc) <= valid_to
                    except Exception as e:
                        result_entry["is_valid"] = f"Error parsing: {str(e)}"
                
                results["found_in"].append(result_entry)
        except Exception as e:
            logger.error(f"Error checking {collection_name}: {str(e)}")
    
    return results

# ============ DASHBOARD ============

@api_router.get("/registered-visitors")
async def get_all_registered_visitors():
    """Get all faculty-registered visitors"""
    visitors = await db.faculty_registered_visitors.find({}, {"_id": 0}).to_list(5000)
    # Add faculty name for each visitor
    for visitor in visitors:
        faculty = await db.faculty.find_one({"faculty_id": visitor["faculty_id"]}, {"_id": 0})
        if faculty:
            visitor["faculty_name"] = faculty["name"]
    return visitors

@api_router.get("/alumni/export")
async def export_alumni_data(date_range: str = "all"):
    """Export alumni data as Excel with optional date filtering"""
    try:
        # Determine date filter based on date_range parameter
        filter_query = {}
        now = datetime.now(timezone.utc)
        
        if date_range == "last_month":
            start_date = (now - timedelta(days=30)).isoformat()
            filter_query = {"created_at": {"$gte": start_date}}
        elif date_range == "last_3_months":
            start_date = (now - timedelta(days=90)).isoformat()
            filter_query = {"created_at": {"$gte": start_date}}
        elif date_range == "last_6_months":
            start_date = (now - timedelta(days=180)).isoformat()
            filter_query = {"created_at": {"$gte": start_date}}
        elif date_range == "last_year":
            start_date = (now - timedelta(days=365)).isoformat()
            filter_query = {"created_at": {"$gte": start_date}}
        elif date_range == "last_2_years":
            start_date = (now - timedelta(days=730)).isoformat()
            filter_query = {"created_at": {"$gte": start_date}}
        
        # Fetch alumni with filter
        alumni_data = await db.alumni.find(filter_query, {"_id": 0}).to_list(5000)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alumni"
        
        # Define headers
        headers = [
            "First Name", "Middle Name", "Last Name", "Email", "Secondary Email",
            "Mobile No", "Secondary Phone", "Enrollment Number", "Degree",
            "Department", "Sub Institute", "Year of Joining", "Year of Passing",
            "Company", "Designation", "LinkedIn Profile", "Address",
            "Whom to Meet", "Selected Person Name", "Selected Person Mobile",
            "Valid Till", "Token"
        ]
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add data
        for row_num, alum in enumerate(alumni_data, 2):
            ws.cell(row=row_num, column=1, value=alum.get("first_name", ""))
            ws.cell(row=row_num, column=2, value=alum.get("middle_name", ""))
            ws.cell(row=row_num, column=3, value=alum.get("last_name", ""))
            ws.cell(row=row_num, column=4, value=alum.get("email", ""))
            ws.cell(row=row_num, column=5, value=alum.get("secondary_email", ""))
            ws.cell(row=row_num, column=6, value=alum.get("mobile_no", ""))
            ws.cell(row=row_num, column=7, value=alum.get("secondary_phone", ""))
            ws.cell(row=row_num, column=8, value=alum.get("enrollment_number", ""))
            ws.cell(row=row_num, column=9, value=alum.get("degree", ""))
            ws.cell(row=row_num, column=10, value=alum.get("department", ""))
            ws.cell(row=row_num, column=11, value=alum.get("sub_institute", ""))
            ws.cell(row=row_num, column=12, value=alum.get("year_of_joining", ""))
            ws.cell(row=row_num, column=13, value=alum.get("year_of_passing", ""))
            ws.cell(row=row_num, column=14, value=alum.get("company", ""))
            ws.cell(row=row_num, column=15, value=alum.get("designation", ""))
            ws.cell(row=row_num, column=16, value=alum.get("linkedin_profile", ""))
            ws.cell(row=row_num, column=17, value=alum.get("address", ""))
            ws.cell(row=row_num, column=18, value=alum.get("whom_to_meet", ""))
            ws.cell(row=row_num, column=19, value=alum.get("selected_person_name", ""))
            ws.cell(row=row_num, column=20, value=alum.get("selected_person_mobile", ""))
            ws.cell(row=row_num, column=21, value=alum.get("valid_till", ""))
            ws.cell(row=row_num, column=22, value=alum.get("token", ""))
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=alumni_{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Error exporting alumni data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting alumni data: {str(e)}")

@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    
    # Count today's entries
    visitors_today = await db.visitors.count_documents({"created_at": {"$gte": today_start}})
    
    # Total counts
    total_students = await db.students.count_documents({})
    total_faculty = await db.faculty.count_documents({})
    total_visitors = await db.visitors.count_documents({})
    total_alumni = await db.alumni.count_documents({})
    total_events = await db.events.count_documents({})
    total_registered_visitors = await db.faculty_registered_visitors.count_documents({})
    
    return {
        "visitors_today": visitors_today,
        "total_students": total_students,
        "total_faculty": total_faculty,
        "total_visitors": total_visitors,
        "total_alumni": total_alumni,
        "total_events": total_events,
        "total_registered_visitors": total_registered_visitors
    }

# ============ DIAGNOSTIC ENDPOINTS ============

@api_router.get("/health/email")
async def check_email_health():
    """Check Gmail SMTP connectivity and configuration"""
    diagnostics = {
        "service": "Gmail SMTP",
        "status": "unknown",
        "server": GMAIL_SMTP_SERVER,
        "port": GMAIL_SMTP_PORT,
        "sender_email": SENDER_EMAIL,
        "issues": []
    }
    
    # Check environment variables
    if not SENDER_EMAIL:
        diagnostics["issues"].append("❌ SENDER_EMAIL not configured in .env")
    
    if not GMAIL_PASSWORD:
        diagnostics["issues"].append("❌ GMAIL_PASSWORD not configured in .env")
        diagnostics["status"] = "failed"
        return diagnostics
    
    # Try to connect to Gmail SMTP
    def test_gmail_connection():
        try:
            with smtplib.SMTP(GMAIL_SMTP_SERVER, GMAIL_SMTP_PORT, timeout=5) as server:
                server.starttls()
                server.login(SENDER_EMAIL, GMAIL_PASSWORD)
                return True, None
        except smtplib.SMTPAuthenticationError as e:
            return False, f"Authentication failed: {str(e)}"
        except TimeoutError:
            return False, f"Connection timeout: Could not reach {GMAIL_SMTP_SERVER}:{GMAIL_SMTP_PORT}"
        except OSError as e:
            return False, f"Network error: {str(e)}"
        except Exception as e:
            return False, f"Unexpected error: {str(e)}"
    
    try:
        success, error = await asyncio.to_thread(test_gmail_connection)
        if success:
            diagnostics["status"] = "✅ OK - Gmail SMTP working"
        else:
            diagnostics["status"] = "❌ FAILED"
            diagnostics["issues"].append(error)
    except Exception as e:
        diagnostics["status"] = "❌ FAILED"
        diagnostics["issues"].append(f"Test error: {str(e)}")
    
    return diagnostics

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()